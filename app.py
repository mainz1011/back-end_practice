# Functions
'''
def greet_user():
    print('Hi there!')
    print('Welcome aboard')


print("Start")
greet_user()
print("Finish")
'''

# Parameters (holders/placeholders that receive and pass info to functions)
'''
def greet_user(name):
    name = "John"            #this line of code isn't needed actually
    print(f'Hi {name}!')
    print('Welcome aboard')


print("Start")
greet_user("John")      # here  "John"  is an argument (the piece of info that's supplied to a function)  
print("Finish")
'''
#utilize the function to make it more efficient
'''
def greet_user(name):          
    print(f'Hi {name}!')
    print('Welcome aboard')


print("Start")
greet_user("John")
greet_user("Marie")        # the function can be reused
print("Finish")
'''
'''
def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard')


print("Start")
greet_user("John", "Smith")        # They're Positional arguments, so position matters;
greet_user("Marie", "Curie")         #while for 'Keyword arguments', position doesn't matter (see the section below)
print("Finish")
'''

# Keyword arguments (position doesn't matter. In certain situations, Keyword arguments can help improve the code readability)
                                              #e.g., calc_cost(total=50, shipping=5, discount=0.2)
  #ALWAYS use the Positional arguments first, and then use Keyword arguments.
'''
def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard')


print("Start")
greet_user(last_name="Smith", first_name="John")    #keyword argument
greet_user("Marie", "Curie")
print("Finish")
'''

# Return statement: how to create functions that return values
  #by default, in Python, all functions return None
'''
def square(number):
    return number * number

print(square(3))
'''

# creating a reusable function
'''
def emoji_converter(message):
    words = message.split(" ")
    emojis = {
        ":D": "😀",
        ":(": "🙁",
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


message = input(">")         # input can come in different forms, e.g., terminal (here in this case), GUI, etc.
print(emoji_converter(message))
'''

# Exceptions
'''
try:
    age = int(input('Age: '))
    print(age)
except ValueError:
    print('Invalid value')
'''

'''
try:
    age = int(input('Age: '))
    income = 20000
    risk = income / age
    print(age)
except ValueError:
    print('Invalid value')
except ZeroDivisionError:
    print('Age cannot be 0.')
'''

# Classes: a class defines a new type, and the new type can create new objects,
  #so an object is an instance of a class.
  #(a class defines the blueprint/template for creating objects.)
  #NOTE: every method in our class should have the 'self' parameter, which should be the 1st parameter of every method
'''
class Point:
    def move(self):
        print("move")
    def draw(self):
        print("draw")

#creating a new object (by typing out the class name and then calling it like a function, see below)
#Point()      #this creates a new object 'Point' and then returns it
#and then we store the new object in a new variable (as the 1st line below shows)
point1 = Point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()
#point2 = Point()     #this object created results in the  AttributeError: 'Point' object has no attribute 'x'
#point2.x = 1          #now, if we assign a value to its x, then the results would show 1 
#print(point2.x)        #so each object is a different instance of our 'Point' class
'''
'''results:  
10
draw
'''

# Constructors
'''
class Point:
    def __init__(self, x, y):     #here a constructor is created to initialize the object 'Point() below'
        self.x = x                 #it's the function/method that gets called when we create a new 'Point()' object
        self.y = y      # 'self' here is a reference to the current object 'Point()', and then we set the x/y attribute to the x/y argument passed to the function
    def move(self):
        print("move")
    def draw(self):
        print("draw")


point = Point(10, 20)   #when we create a new 'Point()' object, 'self' references this object in memory
point.x = 10
#point.x = 11   #here the results would be changed to 11
print(point.x)    #results: 10
'''
#quiz
'''
class Person:
    def __init__(self, name):
        self.name = name   #we set the 'name' attribute of the current 'Person()' object to the 'name' argument passed to this function
    def talk(self):
        print(f"Hi, I am {self.name}")   # self.name  returns the 'name' attribute of the 'Person()' object


john = Person("John Smith")
#print(john.name)   #this would print results: John Smith
john.talk()

bob = Person("Bob Smith")
bob.talk()
'''
'''results:
Hi, I am John Smith
Hi, I am Bob Smith
'''

# Inheritance (a mechanism for using code)
'''
class Mammal:       #here 'Mammal' is a parent class
    def walk(self):
        print("walk")


class Dog(Mammal):
    #pass     #we're telling Python interpreter to just pass this line
    def bark(self):
        print("bark")

class Cat(Mammal):
    def be_annoying(self):
        print("annoying")

dog1 = Dog()
dog1.walk()
dog1.bark()
cat1 = Cat()
cat1.walk()
cat1.be_annoying()
'''
'''results:
walk
bark
walk
annoying
'''

# Modules: a module in Python is a file with some code. We use modules to organize our code into files.
  #see 'converters.py' (in this example, we took 2 functions and put them in a separate module, i.e., 'converters.py')
'''
import converters    #import the entire module 'converters.py'
print(converters.kg_to_lbs(70))
'''
'''
from converters import kg_to_lbs    #import specific functions (or classes) from a module
print(kg_to_lbs(100))
'''
#quiz
'''
from utils import find_max
numbers = [10, 6, 4, 7, 8]   #we first define the list of numbers

#max = find_max(numbers)    #we pass the list as an argument to 'find_max'
#print(max)      #'print' and 'max' are both built-in functions. But here, 'coz of the above line, 'max' is also an integer
print(max(numbers))
'''

# Packages: another way to organize our code. It's a container for multiple modules. It's usually a directory or folder.
  #packages are esp. important when working with Django.
'''
from ecommerce.shipping import calc_shipping, calc_tax
calc_shipping()
calc_tax() 
'''

# (Using a built-in module for) Generating random values
  #NOTE: Built-in modules in Python. Python comes with a standard library that contains several modules for common tasks,
   #e.g., sending emails, working with datetime, encoding values, [generating random values and pwds], and so on.
'''
import random

for i in range(3):     #using the 'range' function to create a 'range' object
    print(random.random())
'''
'''results1: random values between 0-1
0.5359721915107716
0.8936570903449269
0.8846254073065869
'''
'''results2:
0.8859190294416429
0.35275144925974067
0.1954442270130884
'''
'''
import random

for i in range(3):
    print(random.randint(10, 20))
'''
'''results1: random integers between 10-20
10
19
10
'''
'''results2:
18
17
20
'''
#quiz1: randomly pick sb. as the leader
'''
import random

members = ['John', 'Mary', 'Bob', 'Ross']
leader = random.choice(members)
print(leader)
'''
#quiz2: roll & dice
'''
import random


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())
'''

# Files and directories
 #NOTE: 'pathlib' ( https://docs.python.org/3/py-modindex.html ) provides classes that we can use to create objects to work with directories and files.
#from pathlib import Path      #here 'Path' is a class
#p = Path('.')                (#here an instance of the 'Path' class is created)
'''
from pathlib import Path   #Now we need to create a 'path' object to reference a file/directory on our PC.
                            #2 ways to do it: Absolute path (from root of our hard disk, e.g., c:\Program Files\Microsoft...); Relative path
path = Path("ecommerce")
print(path.exists())     #results: True
'''
'''
from pathlib import Path
path = Path("emails")
print(path.mkdir())     # mkdir() created a directory 'emails'. results: None
                          # rmdir() means to remove a directory
'''
'''
from pathlib import Path
path = Path()
print(path.glob('*.py'))      # glob()  allows us to search for files&directories in the current path
                         #glob('*') :look for everything; glob('*.*') : look for all the files;
                         #glob('*.py'): look for all Python files; glob('*.xls'): look for all speadsheets.
                         #and we can iterate/loop through all the generator objects.
'''
'''
from pathlib import Path
path = Path()
for file in path.glob('*.py'):
    print(file)      #results: (vertically list the names of all .Py files in the current directory)
'''

# Pypi and Pip
  #web scraping: build an engine that browses a website and extract info from HTML files.
   #Google uses the same technique to index various websites as it uses several engines (or web crawlers) to continuously browse websites and extract info.
    #selenium (a powerful Python package for browser automation: automate the testing of web apps).


# AUTOMATION WITH PYTHON
# PROJECT1: Excel spreadsheets (1.add a new column; 2.draw a bar chart)
'''
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#the function below takes the name of a file as an input
#def process_workbook(filename):

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['d1']  # or: cell = sheet.cell(1, 4)
    # print(cell.value)     #results would be the value of the specific cell
    # print(sheet.max_row)   #results would be the number of rows
cell.value = 'corrected_price'
for row in range(2, sheet.max_row + 1):  # range(1,4) will generate 1,2,3 (4 won't be included)
    cell = sheet.cell(row, 3)
    # print(cell.value)      #results: (vertically list the value of each row (except row1) at column3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)  # add a new column
    corrected_price_cell.value = corrected_price
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('transactions2.xlsx')
'''

# ML with Python (scikit-learn, the most popular ML library: https://scikit-learn.org/stable/)
  #ML is used for: forecasting stock market trends, robotics, language processing, vision processing, self-driving cars
  #the environment for ML: Jupyter (for inspecting the data and writing our code)
  #ML needs a large amount of data (sometimes even millions of samples)
#ML in action
'''Steps:
1. import the data
2. clean/prepare the data
3. split the data into 2 segments: a general RULE OF THUMB - 1) for training (allocate 70-80% of the data to train the model); 2) for testing (the rest 20-30% is to test the model)
4. create/build a model (involves selecting an algorithm to analyze the data)   # ML algorithms: e.g., decision trees, neural networks (each algorithm has pros&cons in terms of its performance&accuracy)
   #after building the model, we should measure its accuracy. If not accurate enough, we should either fine tune it or build a model using a different algorithm 
5. train the model (so it learns patterns in the data)
6. make predictions
7. evaluate and improve 
'''

#Libraries and tools for ML
 #Libraries:
  # • Numpy: it provides a multi-dimensional array.
  # • Pandas: it's a data analysis library that provides a concept called 'data frame' (a 2D data structure similar to an Excel spreadsheet)
              #the data frame has rows&columns where we can select data in a row or a column or a range of rows&columns
              #Pandas is very popular in ML/DS (data science) projects
  # • MatPlotLib: it's a 2D plotting library for creating graphs on plot.
  # • scikit-learn: it provides all the common algorithms e.g., decision trees, neural networks, and so on


# Jupyter Notebook (on its dashboard http://localhost:8888/tree , the "Files" tab shows my home directory C:\Users\user )
 #from "Files" we create a notebook for Python3 where we can write Python code and execute line by line (& visualize)
  #a Jupyter Notebook "HiNewbie.ipynb" is created for our ML project
   #".ipynb" is similar to ".py" but it includes additional data that Jupyter uses to execute our code
                                     #(".ipynb" includes our source code organized in cells as well as the output for each cell)
                                     #while ".py" only has the source code.

#importing a data set in Jupyter: how to load a data set from a .csv file
 #in "HiNewbie.ipynb", the line   df = pd.read_csv(r"C:\Users\user\PycharmProjects\practice\vgsales.csv")    returns a data frame object which's like an Excel spreadsheet
  #and then we store it as 'df', so we have  df = pd.read_csv(r"C:\Users\user\PycharmProjects\practice\vgsales.csv")
   #and then we type   df  to inspect it (see lines of code below)
   '''
   import pandas as pd
   df = pd.read_csv(r"C:\Users\user\PycharmProjects\practice\vgsales.csv")
   df       #here results would show several first and last rows&columns of vgsales.csv , ending with "16598 rows × 11 columns"
   '''
   #some examples of what Jupyter can do:
    # df.shape          #results: (16598, 11)
    # df.describe()     #results (  .describe()  returns some basic info about each column with numerical values in this data set. )
    # df.values         #results (  .values      returns a 2D array as it has 2 sets of square brackets representing an outer array and an inner array)

#Jupyter shortcuts (when in command mode (blue vertical line), press  h  to check out all shortcuts)
#Jupyter Notebook also has autocompletion and IntelliSense:
 #e.g., by typing  .df  and then press Tab, you can see all the attributes and methods in this object.


# a real ML project (C:\Users\user\PycharmProjects\practice\a_real_ML_project.ipynb)
 #1. import the data set 'music.csv'
 #2. clean/prepare the data
 #3. split the data into 2 separate data sets: 1) 1st & 2nd columns: input set X; 2) 3rd column; output set y.
     #we train a model here: The 3rd column contains the predictions. After that, we give the model a new input set, and we ask the model to predict
 #---learning and predicting
 #4. create a model by using an ML algorithm called Decision Tree which's provided by scikit-learn,
  #code:   from sklearn.tree import DecisionTreeClassifier   (sklearn is a package that comes with the scikit-learn library)
     #sklearn.tree (here 'tree' is a module), 'DecisionTreeClassifier' is a class that implements the Decision Tree algorithm
  #code (importing the 'train_test_split' algorithm):   from sklearn.model_selection import train_test_split
  #code:   model = DecisionTreeClassifier()      #here a model is created
 #5. train the model
  #code:   model.fit(X.values, y.values)    #this method takes the input (X) and output (y) data sets
 #6. make predictions      #before predicting, we temporarily inspect the initial data set (code:  music_data )
  #code:   predictions = model.predict([ [21, 1], [22, 0] ])      #this method takes a 2D array: an outer array and an inner array
                                       #here we pass 2 new input sets to the outer array, so we ask the model to make 2 predictions at the same time
  #code:   predictions      #here we inspect our code again
  #code (showing the predictions):  array(['HipHop', 'Dance'], dtype=object)
 #7. evaluate and improve

#Calculating the accuracy of a model
  #code (example):   from sklearn.metrics import accuracy_score

#Persisting models by storing the model that's been trained
  #code (example):   import joblib
  #                  joblib.dump(model, 'music-recommender.joblib')

#Loading the trained model
'''code (example):   
model = joblib.load('music-recommender.joblib')
predictions = model.predict([ [21, 1], [22, 0] ])
predictions
'''

#Visualizing a Decision Tree (exporting our model in a visual format)
'''code (example):   
from sklearn import tree 

tree.export_graphviz(model, out_file='music-recommender.dot',   
                     feature_names=['age','gender'],   
                     class_names=sorted(y.unique()),   
                     label='all', 
                     rounded=True, 
                     filled=True)
'''
#in CS Code, from Command Palette, select "Graphviz: Open preview to the side": now we can see a binary tree (each node has a max of 2 children)
 #our model generates the rules in the Decision Tree based on the patterns it finds in our data set. As we give our model more data, these rules will change.

# My 1st Django project
